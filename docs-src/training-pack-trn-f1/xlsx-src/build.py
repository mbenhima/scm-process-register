import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import os

TEAL_DEEP = '15423A'
TEAL = '1F6459'
MINT = 'EAF2EF'
ORANGE = 'C2661C'
ORANGE_LIGHT = 'F7E4D2'
INK = '16221F'
MUTED = '5C6D68'
WHITE = 'FFFFFF'
FONT = 'Arial'

thin = Side(style='thin', color='DBE6E3')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_header(cell):
    cell.font = Font(name=FONT, bold=True, color=WHITE, size=11)
    cell.fill = PatternFill('solid', fgColor=TEAL_DEEP)
    cell.alignment = Alignment(vertical='center', wrap_text=True)
    cell.border = border

def style_body(cell, alt=False, wrap=True):
    cell.font = Font(name=FONT, size=10.5, color=INK)
    cell.fill = PatternFill('solid', fgColor=MINT if alt else WHITE)
    cell.alignment = Alignment(vertical='top', wrap_text=wrap)
    cell.border = border

def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

def freeze_header(ws, row=2):
    ws.freeze_panes = f'A{row}'

wb = openpyxl.Workbook()

# =====================================================================
# SHEET 1 — OVERVIEW
# =====================================================================
ws = wb.active
ws.title = 'Overview'
ws.sheet_view.showGridLines = False
set_widths(ws, [22, 90])

ws['A1'] = 'TRN-F1 Training Pack'
ws['A1'].font = Font(name=FONT, bold=True, size=20, color=TEAL_DEEP)
ws['A2'] = 'Change Management Foundations & the journi Platform'
ws['A2'].font = Font(name=FONT, bold=True, size=13, color=TEAL)
ws['A3'] = 'journi Academy — Foundation Level · Prepared for POWERACT Consulting · Confidential'
ws['A3'].font = Font(name=FONT, italic=True, size=10, color=MUTED)
ws.row_dimensions[1].height = 26
ws.merge_cells('A1:B1'); ws.merge_cells('A2:B2'); ws.merge_cells('A3:B3')

rows = [
    ('Training ID', 'TRN-F1'),
    ('Name', 'Change Management Foundations & the journi Platform'),
    ('Level', 'Foundation (LVL-F)'),
    ('Prerequisite', 'None — entry training for the Foundation level'),
    ('Duration', '2 days (4 half-days)'),
    ('Target audience', 'New Change Managers, HR Business Partners, PMO members, People Managers'),
    ('Goals', '1) Explain why ~70% of transformations fail on the human side and how ADKAR, Kotter, Lewin '
              'and Bridges/Kübler-Ross address it. 2) Navigate tenant hierarchy (Group/Organization/Project) '
              'and role-based access. 3) Read the OBS and the Macro Process/SIPOC/RACSI registry. '
              '4) Explain the Assistive/Augmented AI governance model and why Autonomous AI is out of scope.'),
    ('Golden rule', 'Every half-day pairs one quiz with one workshop: content → quiz → workshop → debrief.'),
    ('Training exam', 'EXM-F1 — 40 MCQ + 2 short-answer, 75 min, pass mark 75%'),
    ('Workbook contents', 'Agenda · Quiz Bank · Workshop Checklist · Exam Blueprint · Cohort Scorecard'),
]
r = 5
for k, v in rows:
    ws.cell(row=r, column=1, value=k)
    style_body(ws.cell(row=r, column=1), alt=True, wrap=False)
    ws.cell(row=r, column=1).font = Font(name=FONT, bold=True, size=10.5, color=TEAL_DEEP)
    ws.cell(row=r, column=2, value=v)
    style_body(ws.cell(row=r, column=2))
    ws.row_dimensions[r].height = 32 if len(v) < 90 else 60
    r += 1

# =====================================================================
# SHEET 2 — AGENDA
# =====================================================================
ws = wb.create_sheet('Agenda')
ws.sheet_view.showGridLines = False
headers = ['Half-Day', 'Theme', 'Content (min)', 'Quiz (min)', 'Workshop (min)', 'Total (min)']
set_widths(ws, [12, 42, 13, 11, 15, 12])
for c, h in enumerate(headers, start=1):
    ws.cell(row=1, column=c, value=h)
    style_header(ws.cell(row=1, column=c))
freeze_header(ws)

half_days = [
    ('Day 1 · AM', 'Why Change Management, Why journi', 120, 15, 90),
    ('Day 1 · PM', 'The Process Backbone & Governed AI', 120, 15, 90),
    ('Day 2 · AM', 'Reading the Program Layer', 105, 15, 90),
    ('Day 2 · PM', 'Metrics, Dashboards & the Field Notebook', 90, 15, 90),
]
r = 2
for i, (label, theme, content, quiz, workshop) in enumerate(half_days):
    vals = [label, theme, content, quiz, workshop]
    for c, v in enumerate(vals, start=1):
        cell = ws.cell(row=r, column=c, value=v)
        style_body(cell, alt=(i % 2 == 1), wrap=(c == 2))
    total_cell = ws.cell(row=r, column=6, value=f'=SUM(C{r}:E{r})')
    style_body(total_cell, alt=(i % 2 == 1), wrap=False)
    r += 1
ws.cell(row=r, column=1, value='Total')
ws.cell(row=r, column=1).font = Font(name=FONT, bold=True, size=10.5, color=WHITE)
for c in range(1, 7):
    cell = ws.cell(row=r, column=c)
    cell.fill = PatternFill('solid', fgColor=TEAL_DEEP)
    cell.border = border
    if c >= 3:
        col = get_column_letter(c)
        cell.value = f'=SUM({col}2:{col}{r-1})'
        cell.font = Font(name=FONT, bold=True, size=10.5, color=WHITE)

# =====================================================================
# SHEET 3 — QUIZ BANK
# =====================================================================
ws = wb.create_sheet('Quiz Bank')
ws.sheet_view.showGridLines = False
headers = ['Quiz ID', 'Item #', 'Topic', 'Question', 'Option A', 'Option B', 'Option C', 'Option D', 'Correct']
set_widths(ws, [10, 8, 26, 46, 30, 30, 30, 30, 9])
for c, h in enumerate(headers, start=1):
    ws.cell(row=1, column=c, value=h)
    style_header(ws.cell(row=1, column=c))
freeze_header(ws)

quiz_bank = [
    ('QZ-F1-1', 'Framework definitions + platform navigation', [
        ('What percentage of organizational transformations are estimated to fail primarily due to the human side of change, not the technical/project side?',
         '30%', '50%', '~70%', '90%', 'C'),
        ('Which of the following is NOT one of the four core change-management frameworks referenced in journi?',
         'ADKAR', "Kotter's 8-Step Model", "Lewin's Change Model", 'SWOT Analysis', 'D'),
        ('The Bridges Transition Model is most closely associated with which companion framework in journi\'s Emotional & Transition layer?',
         'Kübler-Ross Change Curve', 'Balanced Scorecard', 'RACI Matrix', 'PMBOK', 'A'),
        ("What is the primary purpose of journi's scope switcher in the top bar?",
         "To change the application's color theme", 'To switch between Group / Organization / Project scope',
         "To switch the logged-in user's role", 'To toggle dark mode', 'B'),
        ("Which languages does journi's multilingual UI natively support?",
         'English, Spanish, German', 'English, French, Arabic', 'English only', 'French, Arabic, Mandarin', 'B'),
        ('ADKAR is best described as a framework that measures change readiness at which level?',
         'Organizational level only', 'Individual level', 'Vendor level', 'Regulatory level', 'B'),
        ("In journi's Sidebar, modules are primarily organized around which structural principle?",
         'Alphabetical order of module names', 'The change-management lifecycle from process registry through sustainment',
         'Random order set by the user', 'Cost of each module', 'B'),
        ('Why does journi teach all four frameworks (ADKAR, Kotter, Lewin, Bridges/Kübler-Ross) together instead of picking one?',
         'Regulatory compliance requires all four',
         'Each framework illuminates a different dimension of change and together they give a fuller picture',
         'journi only supports one framework technically', 'They are identical frameworks with different names', 'B'),
    ]),
    ('QZ-F1-2', 'RACSI role codes + AI tier definitions', [
        ('In a RACSI grid, what does the "C" stand for?', 'Contributor', 'Consulted', 'Controller', 'Compliance', 'B'),
        ("How many Macro Processes make up journi's seeded Macro Process Catalog?", '5', '8', '10', '15', 'C'),
        ('What is the difference between the OBS roster and the RBAC user list?',
         'They are the same thing',
         'OBS models resourcing/reporting structure for the process backbone, while RBAC users control platform login and permissions',
         'OBS is only for Super Admins', 'RBAC users only exist at Group level', 'B'),
        ('An "Assistive" tier AI use case is best described as one where the AI:',
         'Acts fully autonomously with no human review',
         'Only suggests content that a human must review and explicitly accept, edit, or reject',
         'Makes the final decision but logs it for audit', 'Is disabled by default and cannot be activated', 'B'),
        ('Which AI governance tier is explicitly out of scope in journi?', 'Assistive', 'Augmented', 'Autonomous', 'Manual', 'C'),
        ('An End-to-End (E2E) process chain in journi\'s Process Registry is composed of:',
         'A single Macro Process', 'An ordered sequence of Macro Processes', 'A single SIPOC diagram only', 'A list of RACSI roles', 'B'),
        ('What does a SIPOC diagram capture for a given process?',
         'Suppliers, Inputs, Process, Outputs, Customers', 'Sponsors, Impact, Process, Owners, Cost',
         'Stakeholders, Influence, Priority, Outcome, Change', 'Scope, Initiative, Plan, Objective, Charter', 'A'),
        ('What distinguishes "Augmented" tier AI use cases from "Assistive" ones in journi\'s governance model?',
         'Augmented use cases require no human checkpoint at all',
         'Augmented use cases carry a higher degree of AI-driven action within a still-governed human checkpoint',
         'Augmented is a synonym for Assistive', 'Augmented use cases are only available to Super Admins', 'B'),
    ]),
    ('QZ-F1-3', 'Program layer terms', [
        ('What does the Initiative Registry record for each Change Management program?',
         "Only the program's budget",
         'Metadata such as Lewin macro-state, the justified-change pattern, and the Composite Readiness Index',
         "Only the sponsor's name", 'The AI Usage Log', 'B'),
        ('At the Foundation level, what is a learner expected to do with a CM Charter?',
         'Author a new charter from scratch', 'Read an existing charter, not yet edit it', 'Delete a charter', 'Merge two charters', 'B'),
        ('A WBS (Work Breakdown Structure) in journi typically spans which tracks?',
         'Finance, Legal, HR', 'PM, CM, and Framework tracks', 'Sales and Marketing only', 'A single undivided track', 'B'),
        ('What is a "schedule gap" as read on a Gantt view?',
         'A missing task description', "A difference between a task's baseline (planned) date and its actual date",
         'A gap in the RACSI grid', 'An unassigned AI use case', 'B'),
        ('A Phase Gate in journi is best described as:',
         'A cosmetic milestone with no decision attached',
         'A checkpoint where a Joint Decision is recorded to allow (or hold) progression to the next phase',
         'A synonym for a Macro Process', 'An AI-only automated approval step', 'B'),
        ('What is the "justified-change pattern" referenced in the Initiative Registry?',
         'An unexplained status field', 'A requirement that key state changes be accompanied by a written justification',
         'A financial approval workflow', 'A naming convention for projects', 'B'),
        ('Who is expected to be named as "Accountable" for a Phase Gate Joint Decision?',
         'Whoever last edited the record', "A role independent of either input's author",
         'The AI system', 'No one — it is automatic', 'B'),
        ('At Foundation level, what should a learner be able to summarize about a program from its WBS/Gantt and Phase Gate status?',
         'Nothing — this is Advanced-level only', "The program's overall health in a few concise bullet points",
         'The exact source code of the module', 'The personal email of the sponsor', 'B'),
    ]),
    ('QZ-F1-4', 'Dashboard reading + Field Notes scope', [
        ('At the Foundation level, how should a learner interact with the Metrics & Analytics Dashboard?',
         'Freely edit all underlying data', 'Read it for reference only, without editing',
         'Delete stale metrics', 'Export it to a third-party AI tool', 'B'),
        ('What does the ADKAR heatmap on the dashboard visualize?',
         'Server response times', 'Readiness scores across the 5 ADKAR blocks for a cohort or project',
         'Weather patterns affecting go-live', 'AI Usage Log entries only', 'B'),
        ('The adoption curve on the dashboard is used to track:',
         'Software licensing costs', 'The trend of adoption/usage over time for the change',
         'The number of open support tickets', 'The org chart depth', 'B'),
        ('Field Notes in journi are best described as:',
         'A structured system-of-record replacing the Risk Register',
         "A lightweight practitioner scratchpad for observations that don't yet belong in a structured module",
         'A mandatory daily report to the Super Admin', 'An AI-generated audit log', 'B'),
        ("Why shouldn't a Field Note be used as a substitute for a structured record (e.g., a Risk Register entry)?",
         'Field Notes are automatically deleted after 24 hours',
         'Structured modules provide the scoring, ownership, and workflow that a raw scratchpad entry lacks',
         'Field Notes cannot contain text', 'There is no difference — either works equally well', 'B'),
        ('Which of the following is an appropriate use of a Field Note?',
         'Logging a formal mitigation action with an owner and due date',
         'Capturing a quick observation from a review to revisit or escalate later',
         'Recording a Phase Gate Joint Decision', "Setting a user's RBAC role", 'B'),
        ("A project's ADKAR heatmap shows its weakest block. What should a Foundation-level learner be able to do with that information?",
         "Immediately override the organization's AI governance tier",
         'Identify it as the area most needing intervention or follow-up',
         'Ignore it since it is reference-only', 'Delete the block from the dashboard', 'B'),
        ('What is the relationship between Field Notes and the multilingual UI covered earlier in the training?',
         'Field Notes are only available in English',
         "Field Notes, like the rest of journi's UI, are usable in EN/FR/AR per the practitioner's configured language",
         'Field Notes require a separate language license', 'There is no relationship', 'B'),
    ]),
]

r = 2
for quiz_id, topic, items in quiz_bank:
    for i, (q, a, b, c, d, correct) in enumerate(items, start=1):
        row_vals = [quiz_id, i, topic, q, a, b, c, d, correct]
        alt = (r % 2 == 0)
        for col, v in enumerate(row_vals, start=1):
            cell = ws.cell(row=r, column=col, value=v)
            style_body(cell, alt=alt, wrap=(col in (3, 4, 5, 6, 7, 8)))
        ws.cell(row=r, column=9).alignment = Alignment(horizontal='center', vertical='top')
        ws.cell(row=r, column=9).font = Font(name=FONT, bold=True, size=10.5, color=ORANGE)
        ws.row_dimensions[r].height = 46
        r += 1

# Summary block: item count per quiz via COUNTIF (sanity check formula)
last_data_row = r - 1
r += 2
ws.cell(row=r, column=1, value='Sanity Check — items per quiz (should be 8)')
ws.cell(row=r, column=1).font = Font(name=FONT, bold=True, size=10.5, color=TEAL_DEEP)
r += 1
for quiz_id, topic, items in quiz_bank:
    ws.cell(row=r, column=1, value=quiz_id)
    style_body(ws.cell(row=r, column=1), wrap=False)
    ws.cell(row=r, column=2, value=f'=COUNTIF(A2:A{last_data_row},A{r})')
    style_body(ws.cell(row=r, column=2), wrap=False)
    r += 1

# =====================================================================
# SHEET 4 — WORKSHOP CHECKLIST
# =====================================================================
ws = wb.create_sheet('Workshop Checklist')
ws.sheet_view.showGridLines = False
headers = ['Workshop ID', 'Workshop Name', 'Step #', 'Facilitation Step', 'Completed (Y/N)']
set_widths(ws, [13, 26, 8, 60, 16])
for c, h in enumerate(headers, start=1):
    ws.cell(row=1, column=c, value=h)
    style_header(ws.cell(row=1, column=c))
freeze_header(ws)

workshops = [
    ('WS-F1-1', 'Stand Up a Tenant', [
        'Create a Group and an Organization for a new manufacturing client scenario.',
        'Register a Main Project and a linked Change Management Project.',
        "Set the Organization's Default Language and explain the precedence rule to a neighbor.",
        "Review the M2 user list and identify each seeded role's scope.",
    ]),
    ('WS-F1-2', 'Read the Process Backbone', [
        'Build a 4-role OBS roster with a reporting chain for the scenario project.',
        'Trace one End-to-End process chain through its ordered Macro Processes.',
        'Review 3 AI Use Case Library entries and classify tier + human checkpoint for each.',
        'Identify which module each of the 3 use cases plugs into.',
    ]),
    ('WS-F1-3', "Read a Program's Status", [
        'Open a seeded Initiative Registry entry and its CM Charter; identify the Lewin macro-state and top charter action.',
        'Open the linked WBS/Gantt view and identify 2 tasks with a schedule gap.',
        'Identify the current Phase Gate status and who owns the next decision.',
        "Summarize the program's health in 3 bullet points, as if briefing a manager who has 60 seconds.",
    ]),
    ('WS-F1-4', 'Read a Dashboard, Log a Field Note', [
        "Open the Metrics & Analytics Dashboard for the seeded project and identify the ADKAR heatmap's weakest block.",
        "Identify the adoption curve's current trend.",
        'Add a Field Note capturing an observation from the review (not a structured record).',
        'Explain, in writing, why that observation belongs in Field Notes and not the Risk Register.',
    ]),
]

r = 2
first_data_row = r
for ws_id, ws_name, steps in workshops:
    for i, step in enumerate(steps, start=1):
        alt = (r % 2 == 0)
        ws.cell(row=r, column=1, value=ws_id); style_body(ws.cell(row=r, column=1), alt=alt, wrap=False)
        ws.cell(row=r, column=2, value=ws_name); style_body(ws.cell(row=r, column=2), alt=alt)
        ws.cell(row=r, column=3, value=i); style_body(ws.cell(row=r, column=3), alt=alt, wrap=False)
        ws.cell(row=r, column=3).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=4, value=step); style_body(ws.cell(row=r, column=4), alt=alt)
        cell5 = ws.cell(row=r, column=5, value=None); style_body(cell5, alt=alt, wrap=False)
        cell5.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[r].height = 30
        r += 1
last_data_row = r - 1

dv = DataValidation(type='list', formula1='"Y,N"', allow_blank=True)
ws.add_data_validation(dv)
dv.add(f'E{first_data_row}:E{last_data_row}')

r += 2
ws.cell(row=r, column=1, value='Legend: mark each step Y once the participant/cohort has completed it (pick from the dropdown). Completion % below updates automatically.')
ws.cell(row=r, column=1).font = Font(name=FONT, italic=True, size=9.5, color=MUTED)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
r += 2
ws.cell(row=r, column=1, value='Completion Summary')
ws.cell(row=r, column=1).font = Font(name=FONT, bold=True, size=10.5, color=TEAL_DEEP)
r += 1
summary_header_row = r
for c, h in enumerate(['Workshop ID', 'Steps', 'Completed', '% Complete'], start=1):
    ws.cell(row=r, column=c, value=h); style_header(ws.cell(row=r, column=c))
r += 1
for ws_id, ws_name, steps in workshops:
    ws.cell(row=r, column=1, value=ws_id); style_body(ws.cell(row=r, column=1), wrap=False)
    ws.cell(row=r, column=2, value=f'=COUNTIF($A${first_data_row}:$A${last_data_row},A{r})')
    style_body(ws.cell(row=r, column=2), wrap=False)
    ws.cell(row=r, column=3, value=f'=COUNTIFS($A${first_data_row}:$A${last_data_row},A{r},$E${first_data_row}:$E${last_data_row},"Y")')
    style_body(ws.cell(row=r, column=3), wrap=False)
    pct_cell = ws.cell(row=r, column=4, value=f'=IF(B{r}=0,0,C{r}/B{r})')
    style_body(pct_cell, wrap=False)
    pct_cell.number_format = '0%'
    r += 1

# =====================================================================
# SHEET 5 — EXAM BLUEPRINT
# =====================================================================
ws = wb.create_sheet('Exam Blueprint')
ws.sheet_view.showGridLines = False
set_widths(ws, [24, 76])

ws['A1'] = 'EXM-F1 — Training Exam Blueprint'
ws['A1'].font = Font(name=FONT, bold=True, size=16, color=TEAL_DEEP)
ws.merge_cells('A1:B1')

rows = [
    ('Format', '40 MCQ + 2 short-answer'),
    ('Duration', '75 minutes'),
    ('Pass mark', '75%'),
    ('Retake policy', 'One free retake after a 48-hour review period'),
    ('Scope', 'Drawn from all 4 half-day quizzes: framework definitions, RACSI + AI tiers, program-layer terms, dashboard reading + Field Notes scope'),
    ('Short-answer focus', 'Participants justify a governance decision in writing (e.g. why a Phase Gate needs an independent Accountable role)'),
    ('Gate', 'Passing EXM-F1 is one of three training exams required before the Foundation Level Exam (EXM-LVL-F)'),
]
r = 3
for k, v in rows:
    ws.cell(row=r, column=1, value=k)
    style_body(ws.cell(row=r, column=1), alt=True, wrap=False)
    ws.cell(row=r, column=1).font = Font(name=FONT, bold=True, size=10.5, color=TEAL_DEEP)
    ws.cell(row=r, column=2, value=v)
    style_body(ws.cell(row=r, column=2))
    ws.row_dimensions[r].height = 30 if len(v) < 80 else 46
    r += 1

r += 1
ws.cell(row=r, column=1, value='Coverage by Half-Day Topic')
ws.cell(row=r, column=1).font = Font(name=FONT, bold=True, size=12, color=TEAL_DEEP)
r += 1

total_mcq_row = r
ws.cell(row=r, column=1, value='Total MCQ items (EXM-F1)')
style_body(ws.cell(row=r, column=1), alt=True, wrap=False)
mcq_input_cell = ws.cell(row=r, column=2, value=40)
style_body(mcq_input_cell, wrap=False)
mcq_input_cell.font = Font(name=FONT, size=10.5, color='0000FF')
r += 1
topic_count_row = r
ws.cell(row=r, column=1, value='Half-day topics')
style_body(ws.cell(row=r, column=1), alt=True, wrap=False)
topics_input_cell = ws.cell(row=r, column=2, value=4)
style_body(topics_input_cell, wrap=False)
topics_input_cell.font = Font(name=FONT, size=10.5, color='0000FF')
r += 2

cov_header_row = r
for c, h in enumerate(['Half-Day Topic', 'Approx. MCQ Share'], start=1):
    ws.cell(row=r, column=c, value=h); style_header(ws.cell(row=r, column=c))
r += 1
topics = ['Framework definitions + platform navigation', 'RACSI role codes + AI tier definitions',
          'Program layer terms', 'Dashboard reading + Field Notes scope']
for t in topics:
    ws.cell(row=r, column=1, value=t); style_body(ws.cell(row=r, column=1), wrap=True)
    share_cell = ws.cell(row=r, column=2, value=f'=$B${total_mcq_row}/$B${topic_count_row}')
    style_body(share_cell, wrap=False)
    share_cell.number_format = '0 "MCQ"'
    r += 1

# =====================================================================
# SHEET 6 — COHORT SCORECARD (fill-in template)
# =====================================================================
ws = wb.create_sheet('Cohort Scorecard')
ws.sheet_view.showGridLines = False
headers = ['Participant', 'QZ-F1-1 (/8)', 'QZ-F1-2 (/8)', 'QZ-F1-3 (/8)', 'QZ-F1-4 (/8)',
           'Quiz Total (/32)', 'EXM-F1 MCQ (/40)', 'EXM-F1 Short-Answer (/2)', 'EXM-F1 %', 'Result']
set_widths(ws, [20, 12, 12, 12, 12, 13, 14, 16, 10, 10])
for c, h in enumerate(headers, start=1):
    ws.cell(row=1, column=c, value=h)
    style_header(ws.cell(row=1, column=c))
freeze_header(ws)

example_row = ['Jane Doe (example)', 7, 6, 8, 7, None, 36, 2, None, None]
r = 2
for c, v in enumerate(example_row, start=1):
    cell = ws.cell(row=r, column=c, value=v)
    style_body(cell, alt=False, wrap=False)
    cell.font = Font(name=FONT, italic=True, size=10.5, color=MUTED)
ws.cell(row=r, column=6, value=f'=SUM(B{r}:E{r})')
ws.cell(row=r, column=6).font = Font(name=FONT, italic=True, size=10.5, color=MUTED)
ws.cell(row=r, column=9, value=f'=(G{r}+H{r})/(40+2)')
ws.cell(row=r, column=9).number_format = '0%'
ws.cell(row=r, column=9).font = Font(name=FONT, italic=True, size=10.5, color=MUTED)
ws.cell(row=r, column=10, value=f'=IF(I{r}>=0.75,"Pass","Retake")')
ws.cell(row=r, column=10).font = Font(name=FONT, italic=True, size=10.5, color=MUTED)

for r in range(3, 15):
    for c in range(1, 11):
        cell = ws.cell(row=r, column=c)
        style_body(cell, alt=(r % 2 == 0), wrap=False)
    ws.cell(row=r, column=6, value=f'=IF(COUNT(B{r}:E{r})=0,"",SUM(B{r}:E{r}))')
    ws.cell(row=r, column=9, value=f'=IF(OR(G{r}="",H{r}=""),"",(G{r}+H{r})/(40+2))')
    ws.cell(row=r, column=9).number_format = '0%'
    ws.cell(row=r, column=10, value=f'=IF(I{r}="","",IF(I{r}>=0.75,"Pass","Retake"))')

r = 17
ws.cell(row=r, column=1, value='Legend: fill in the blue-outlined columns (quiz scores, exam MCQ score, exam short-answer score out of 2) for each participant. Quiz Total, EXM-F1 %, and Result are computed automatically.')
ws.cell(row=r, column=1).font = Font(name=FONT, italic=True, size=9.5, color=MUTED)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)

out = os.path.join(os.path.dirname(__file__), '..', 'TRN-F1_Training_Pack.xlsx')
wb.save(out)
print('wrote', out)
